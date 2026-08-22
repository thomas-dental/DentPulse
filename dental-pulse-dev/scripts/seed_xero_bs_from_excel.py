import openpyxl
from datetime import datetime

ORG = "aa4580ee-c27a-4967-8a35-7233b8611928"
PI = "57f7883f-1628-450d-8266-a3a45cfc983d"
TENANT = "3a4da8ad-7f13-4526-8403-bf61b86ab13b"
ACCOUNTS = {
    "8d2975b7-af39-4346-b7e1-e0242954d9ac": "Dentally Control Account",
    "33a392e8-fca1-4bf0-934a-baef68ab18a6": "Petty Cash",
    "a3bfc694-659a-4dea-91cb-28903e24f1b4": "SOUTH ST DENTAL - CREDIT CARD",
    "21c0c645-4977-445b-b714-6afc9c7307c5": "SOUTH ST DENTAL/AT X",
    "b011f75d-d954-4870-adfc-549d34bfd339": "SOUTH STREET DENTAL",
}

wb = openpyxl.load_workbook(
    r"c:\Users\Chirag\Downloads\The_South_Street_Dental_Practice_-_Monthly_comparison (1).xlsx",
    data_only=True,
)
ws = wb["Balance Sheet"]
end_dates = []
for c in range(3, ws.max_column + 1):
    v = ws.cell(5, c).value
    if not v:
        continue
    s = str(v).strip().replace("Sept", "Sep")
    dt = datetime.strptime(s, "%d %b %Y")
    end_dates.append(dt.strftime("%Y-%m-%d"))

# Only rows under "Cash at bank and in hand" (before "Total Cash at bank and in hand")
CASH_ROW_START = 21
CASH_ROW_END = 26

name_to_id = {v: k for k, v in ACCOUNTS.items()}
rows = []
for r in range(CASH_ROW_START, CASH_ROW_END):
    name = ws.cell(r, 2).value
    if name not in name_to_id:
        continue
    aid = name_to_id[name]
    for i, ed in enumerate(end_dates):
        amt = ws.cell(r, 3 + i).value or 0
        fy = int(ed[:4])
        fm = int(ed[5:7])
        fd = f"{ed[:7]}-01"
        rows.append((ORG, PI, TENANT, fd, ed, fy, fm, aid, float(amt)))

lines = [
    f"('{a}','{b}','{c}','{d}','{e}',{fy},{fm},'Cash at bank and in hand','{g}',{amt})"
    for a, b, c, d, e, fy, fm, g, amt in rows
]

sql = (
    f"DELETE FROM xero_balance_sheet WHERE organization_id='{ORG}';\n"
    "INSERT INTO xero_balance_sheet "
    "(organization_id, platform_integration_id, xero_tenant_id, from_date, to_date, "
    "period_year, period_month, section, xero_account_id, amount) VALUES\n"
    + ",\n".join(lines)
    + ";\n"
)
out = r"d:\DentPulse\Project-GIT\Enterprise\dental-pulse-dev\seed_bs.sql"
with open(out, "w", encoding="utf-8") as f:
    f.write(sql)
print(f"Wrote {len(rows)} rows to {out}")
